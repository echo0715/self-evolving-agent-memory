"""SpreadsheetBench adapter tests. No network and no model: the LLM is faked.

Workbooks are real, though, and that is deliberate. The three things that decide
a score on this benchmark are all invisible to a mocked evaluator:

  * upstream's cell comparator, whose numeric rule is a `round(v, 2)`
    *quantization* rather than a tolerance, and which fails a type mismatch;
  * the openpyxl formula trap -- correct arithmetic written as `=SUM(...)` reads
    back as `None` and scores zero;
  * the three-test-case protocol, which is the entire reason an agent cannot
    read the preview and write literal answers.

So the fixtures below build actual .xlsx files and the scoring tests run the
real `score_task`.
"""

from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

from memsys.adapters.spreadsheetbench import (
    EMPTY_MEMORY,
    AgentConfig,
    SpreadsheetAgent,
    TaskSpec,
    build_system_prompt,
    compare_cell_value,
    compare_workbooks,
    find_test_cases,
    generate_cell_names,
    load_manifest,
    parse_action,
    run_generated_code,
    run_task,
    score_task,
)

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    HAS_OPENPYXL = False

needs_openpyxl = unittest.skipUnless(HAS_OPENPYXL, "openpyxl not installed")

TASK = TaskSpec(
    task_id="13-1",
    split="test",
    instruction="Put the sum of column A into B1.",
    instruction_type="Cell-Level Manipulation",
    spreadsheet_path="spreadsheet/13-1",
    answer_position="B1",
    answer_sheet="Sheet1",
)


# ------------------------------------------------------------------ fixtures
def write_book(path: str, rows: list[list], sheet: str = "Sheet1") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def make_task_dir(root: str, task_id: str, n_cases: int = 3,
                  suffix: tuple[str, str] = ("_input.xlsx", "_answer.xlsx")) -> str:
    """Three cases of `[a, b, c]` in column A with their sum expected in B1."""
    task_dir = os.path.join(root, "spreadsheet", task_id)
    os.makedirs(task_dir, exist_ok=True)
    for i in range(1, n_cases + 1):
        values = [i * 1, i * 2, i * 3]
        write_book(os.path.join(task_dir, f"{i}_{task_id}{suffix[0]}"), [[v] for v in values])
        write_book(os.path.join(task_dir, f"{i}_{task_id}{suffix[1]}"),
                   [[values[0], sum(values)], [values[1]], [values[2]]])
    return task_dir


#: A correct solution: reads the input, computes, writes a literal.
GOOD_SOLUTION = """import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb.active
ws["B1"] = sum(c[0].value for c in ws.iter_rows(min_col=1, max_col=1))
wb.save(OUTPUT_PATH)
"""

#: The failure the three-case protocol exists to catch: right for case 1 only.
HARDCODED_SOLUTION = """import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
wb.active["B1"] = 6
wb.save(OUTPUT_PATH)
"""

#: Correct arithmetic, zero score: openpyxl writes no cached value for a formula.
FORMULA_SOLUTION = """import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
wb.active["B1"] = "=SUM(A1:A3)"
wb.save(OUTPUT_PATH)
"""


class FakeClient:
    """A reply is either a fixed string or `f(messages) -> str`.

    The callable form exists because a solution.py must declare INPUT_PATH and
    OUTPUT_PATH itself -- nothing injects them while the agent is running, only
    `run_generated_code` does, and only for cases 2 and 3. A fake that emitted
    a script without them would test a protocol the real agent never follows.
    """

    def __init__(self, responses):
        self.responses = list(responses)
        self.calls: list[list[dict]] = []
        self.last_stop = None
        outer = self

        class _Completions:
            def create(self, model, messages, temperature, max_tokens, stop=None):
                outer.calls.append([dict(m) for m in messages])
                outer.last_stop = stop
                reply = outer.responses.pop(0) if outer.responses else "DONE"
                text = reply(messages) if callable(reply) else reply
                return SimpleNamespace(
                    choices=[SimpleNamespace(message=SimpleNamespace(content=text))],
                    usage=SimpleNamespace(prompt_tokens=10, completion_tokens=5),
                )

        self.chat = SimpleNamespace(completions=_Completions())


def paths_from_prompt(messages) -> tuple[str, str]:
    """Read the two paths out of the task prompt, as the real agent must."""
    text = messages[1]["content"]
    fields = {}
    for header in ("# Input file", "# Output file"):
        chunk = text.split(header, 1)[1].strip()
        fields[header] = chunk.splitlines()[0].strip()
    return fields["# Input file"], fields["# Output file"]


def solve_with(code: str) -> list:
    """The two replies that write and run a solution, then stop."""

    def write(messages):
        inp, outp = paths_from_prompt(messages)
        header = f"INPUT_PATH  = {inp!r}\nOUTPUT_PATH = {outp!r}\n"
        return f"I will write the script.\n```python\n{header}{code}```"

    return [write, "Now run it.\n```bash\npython solution.py\n```", "DONE"]


# --------------------------------------------------------------------- prompt
class PromptTest(unittest.TestCase):
    def test_memory_block_is_injected_under_the_header(self):
        prompt = build_system_prompt("- always reopen the workbook")
        self.assertIn("- always reopen the workbook", prompt)
        self.assertNotIn(EMPTY_MEMORY, prompt)

    def test_empty_memory_still_fills_the_section(self):
        # Dropping the section for `none` would make the no-memory arm a
        # different scaffold rather than the same scaffold with empty memory.
        prompt = build_system_prompt("")
        self.assertIn(EMPTY_MEMORY.strip(), prompt)

    def test_critical_rules_survive_in_every_arm(self):
        for block in ("", "some memory"):
            self.assertIn("NEVER write Excel formulas", build_system_prompt(block))


class ParseActionTest(unittest.TestCase):
    def test_python_block_writes_solution_py(self):
        self.assertEqual(parse_action("plan\n```python\nx = 1\n```"),
                         ("write", "solution.py", "x = 1\n"))

    def test_python_block_may_name_its_path(self):
        kind, path, _ = parse_action("```python helper.py\nx = 1\n```")
        self.assertEqual((kind, path), ("write", "helper.py"))

    def test_bash_block_runs(self):
        self.assertEqual(parse_action("```bash\nls -la\n```"), ("bash", "", "ls -la"))

    def test_sh_and_shell_are_bash(self):
        for lang in ("sh", "shell"):
            self.assertEqual(parse_action(f"```{lang}\nls\n```")[0], "bash")

    def test_first_block_wins(self):
        # A model emitting several has invented the shell output between them,
        # so every later block is written against a state that never existed.
        text = "```bash\nfirst\n```\nOutput:\n```\nok\n```\n```bash\nsecond\n```"
        self.assertEqual(parse_action(text), ("bash", "", "first"))

    def test_prose_yields_nothing(self):
        self.assertIsNone(parse_action("The output file is correct. DONE"))


# ----------------------------------------------------------------- evaluator
class CellComparatorTest(unittest.TestCase):
    """Upstream's rules, which are stricter and stranger than they look."""

    def test_numbers_are_quantized_to_two_decimals_not_tolerated(self):
        self.assertTrue(compare_cell_value(1.234, 1.2349))   # both round to 1.23
        self.assertFalse(compare_cell_value(1.234, 1.236))   # 1.23 vs 1.24

    def test_numeric_strings_compare_as_numbers(self):
        self.assertTrue(compare_cell_value("5", 5))
        self.assertTrue(compare_cell_value("5.0", 5))

    def test_none_and_empty_string_are_equal(self):
        self.assertTrue(compare_cell_value(None, ""))
        self.assertTrue(compare_cell_value("", None))

    def test_type_mismatch_fails(self):
        self.assertFalse(compare_cell_value("abc", 1))

    def test_bool_promotes_to_number(self):
        self.assertTrue(compare_cell_value(True, 1))


class RangeTest(unittest.TestCase):
    def test_single_cell(self):
        self.assertEqual(generate_cell_names("B2"), ["B2"])

    def test_rectangular_range_is_column_major(self):
        self.assertEqual(generate_cell_names("A1:B2"), ["A1", "A2", "B1", "B2"])

    def test_multi_letter_columns(self):
        self.assertEqual(generate_cell_names("Z1:AA1"), ["Z1", "AA1"])


@needs_openpyxl
class CompareWorkbooksTest(unittest.TestCase):
    def test_identical_books_match_with_full_cell_rate(self):
        with tempfile.TemporaryDirectory() as d:
            a = os.path.join(d, "a.xlsx")
            write_book(a, [[1, 6], [2], [3]])
            ok, msg, rate = compare_workbooks(a, a, "Sheet1!B1")
            self.assertTrue(ok, msg)
            self.assertEqual(rate, 1.0)

    def test_missing_output_file_is_a_failure_not_an_exception(self):
        with tempfile.TemporaryDirectory() as d:
            a = os.path.join(d, "a.xlsx")
            write_book(a, [[1]])
            ok, msg, rate = compare_workbooks(a, os.path.join(d, "nope.xlsx"), "Sheet1!A1")
            self.assertFalse(ok)
            self.assertEqual(msg, "file not exist")

    def test_missing_sheet_is_a_failure_not_an_exception(self):
        with tempfile.TemporaryDirectory() as d:
            a, b = os.path.join(d, "a.xlsx"), os.path.join(d, "b.xlsx")
            write_book(a, [[1]], sheet="Data")
            write_book(b, [[1]], sheet="Other")
            ok, msg, _ = compare_workbooks(a, b, "Data!A1")
            self.assertFalse(ok)
            self.assertIn("worksheet not found", msg)

    def test_answer_position_may_list_several_quoted_ranges(self):
        with tempfile.TemporaryDirectory() as d:
            a = os.path.join(d, "a.xlsx")
            write_book(a, [[1, 2], [3, 4]], sheet="My Sheet")
            ok, msg, _ = compare_workbooks(a, a, "'My Sheet'!A1:A2,'My Sheet'!B1")
            self.assertTrue(ok, msg)

    def test_cell_match_rate_is_partial_when_one_cell_is_wrong(self):
        with tempfile.TemporaryDirectory() as d:
            a, b = os.path.join(d, "a.xlsx"), os.path.join(d, "b.xlsx")
            write_book(a, [[1, 2, 3, 4]])
            write_book(b, [[1, 2, 3, 99]])
            ok, _, rate = compare_workbooks(a, b, "Sheet1!A1:D1")
            self.assertFalse(ok)
            self.assertEqual(rate, 0.75)


# ------------------------------------------------------------- test cases
@needs_openpyxl
class FindTestCasesTest(unittest.TestCase):
    def test_912_input_answer_naming(self):
        with tempfile.TemporaryDirectory() as d:
            task_dir = make_task_dir(d, "13-1")
            self.assertEqual([c[0] for c in find_test_cases(task_dir)], ["1", "2", "3"])

    def test_verified_400_init_golden_naming(self):
        with tempfile.TemporaryDirectory() as d:
            task_dir = make_task_dir(d, "13-1", n_cases=1,
                                     suffix=("_init.xlsx", "_golden.xlsx"))
            self.assertEqual(len(find_test_cases(task_dir)), 1)

    def test_bare_initial_golden_fallback(self):
        with tempfile.TemporaryDirectory() as d:
            task_dir = os.path.join(d, "spreadsheet", "58109")
            os.makedirs(task_dir)
            write_book(os.path.join(task_dir, "initial.xlsx"), [[1]])
            write_book(os.path.join(task_dir, "golden.xlsx"), [[1, 1]])
            self.assertEqual([c[0] for c in find_test_cases(task_dir)], ["1"])

    def test_an_input_without_its_answer_is_not_a_case(self):
        with tempfile.TemporaryDirectory() as d:
            task_dir = os.path.join(d, "spreadsheet", "x")
            os.makedirs(task_dir)
            write_book(os.path.join(task_dir, "1_x_input.xlsx"), [[1]])
            self.assertEqual(find_test_cases(task_dir), [])

    def test_cases_sort_numerically_not_lexically(self):
        with tempfile.TemporaryDirectory() as d:
            task_dir = os.path.join(d, "spreadsheet", "x")
            os.makedirs(task_dir)
            for i in (1, 2, 10):
                write_book(os.path.join(task_dir, f"{i}_x_input.xlsx"), [[1]])
                write_book(os.path.join(task_dir, f"{i}_x_answer.xlsx"), [[1]])
            self.assertEqual([c[0] for c in find_test_cases(task_dir)], ["1", "2", "10"])


# ------------------------------------------------------------------ executor
@needs_openpyxl
class RunGeneratedCodeTest(unittest.TestCase):
    def test_paths_are_overridden_even_when_the_agent_declared_its_own(self):
        # Left in place, the agent's own INPUT_PATH still points at case 1, and
        # a solution that ignored the override would re-score case 1 three times
        # and look like it generalised.
        with tempfile.TemporaryDirectory() as d:
            src, dst = os.path.join(d, "in.xlsx"), os.path.join(d, "out.xlsx")
            write_book(src, [[2], [4]])
            code = 'INPUT_PATH = "/nope/a.xlsx"\nOUTPUT_PATH = "/nope/b.xlsx"\n' + GOOD_SOLUTION
            ok, err = run_generated_code(code, src, dst)
            self.assertTrue(ok, err)
            wb = openpyxl.load_workbook(dst)
            self.assertEqual(wb.active["B1"].value, 6)
            wb.close()

    def test_a_raising_script_reports_the_traceback(self):
        with tempfile.TemporaryDirectory() as d:
            src, dst = os.path.join(d, "in.xlsx"), os.path.join(d, "out.xlsx")
            write_book(src, [[1]])
            ok, err = run_generated_code("raise ValueError('boom')", src, dst)
            self.assertFalse(ok)
            self.assertIn("boom", err)

    def test_a_script_that_writes_nothing_is_not_a_success(self):
        with tempfile.TemporaryDirectory() as d:
            src, dst = os.path.join(d, "in.xlsx"), os.path.join(d, "out.xlsx")
            write_book(src, [[1]])
            ok, err = run_generated_code("pass", src, dst)
            self.assertFalse(ok)
            self.assertIn("not created", err)


# ------------------------------------------------------------------ scoring
@needs_openpyxl
class ScoringTest(unittest.TestCase):
    """The three-case protocol, which is what the benchmark is actually for."""

    def _score(self, solution: str, case1_from_solution: bool = True):
        d = tempfile.mkdtemp()
        self.addCleanup(lambda: __import__("shutil").rmtree(d, ignore_errors=True))
        task_dir = make_task_dir(d, "13-1")
        cases = find_test_cases(task_dir)
        out_dir = os.path.join(d, "out")
        os.makedirs(out_dir)
        solution_path = os.path.join(out_dir, "solution.py")
        Path(solution_path).write_text(solution, encoding="utf-8")
        if case1_from_solution:
            # Case 1's output is whatever the agent produced during the episode;
            # here we stand in for that by running the same script.
            run_generated_code(solution, cases[0][1], os.path.join(out_dir, "1_pred.xlsx"))
        return score_task(TASK, cases, out_dir, solution_path)

    def test_a_generalising_solution_scores_1_0(self):
        r = self._score(GOOD_SOLUTION)
        self.assertTrue(r["success"])
        self.assertEqual((r["n_pass"], r["n_cases"], r["score"]), (3, 3, 1.0))

    def test_a_hardcoded_solution_passes_case_1_and_fails_the_rest(self):
        r = self._score(HARDCODED_SOLUTION)
        self.assertFalse(r["success"])
        self.assertEqual(r["n_pass"], 1)
        self.assertAlmostEqual(r["score"], 1 / 3)
        self.assertIn("eval-mismatch", r["fail_reason"])

    def test_a_formula_scores_zero_despite_correct_arithmetic(self):
        # openpyxl writes no cached value, so the evaluator reads None.
        r = self._score(FORMULA_SOLUTION)
        self.assertFalse(r["success"])
        self.assertEqual(r["n_pass"], 0)

    def test_no_solution_py_still_scores_the_case_the_agent_produced(self):
        d = tempfile.mkdtemp()
        self.addCleanup(lambda: __import__("shutil").rmtree(d, ignore_errors=True))
        cases = find_test_cases(make_task_dir(d, "13-1"))
        out_dir = os.path.join(d, "out")
        os.makedirs(out_dir)
        run_generated_code(GOOD_SOLUTION, cases[0][1], os.path.join(out_dir, "1_pred.xlsx"))
        r = score_task(TASK, cases, out_dir, os.path.join(out_dir, "missing.py"))
        self.assertEqual((r["n_pass"], r["score"]), (1, 1 / 3))
        self.assertFalse(r["success"])
        self.assertEqual(r["fail_reason"], "no-solution-py-for-other-cases")

    def test_a_solution_that_crashes_on_later_cases_is_an_exec_error(self):
        # Case 1's output is whatever the agent produced in the loop, so a
        # solution.py that cannot run at all still scores case 1 -- the exec
        # error only appears when it is re-applied.
        d = tempfile.mkdtemp()
        self.addCleanup(lambda: __import__("shutil").rmtree(d, ignore_errors=True))
        cases = find_test_cases(make_task_dir(d, "13-1"))
        out_dir = os.path.join(d, "out")
        os.makedirs(out_dir)
        run_generated_code(GOOD_SOLUTION, cases[0][1], os.path.join(out_dir, "1_pred.xlsx"))
        solution_path = os.path.join(out_dir, "solution.py")
        Path(solution_path).write_text("raise ValueError('boom')", encoding="utf-8")
        r = score_task(TASK, cases, out_dir, solution_path)
        self.assertEqual(r["n_pass"], 1)
        self.assertIn("exec-error", r["fail_reason"])
        self.assertIn("boom", r["fail_reason"])

    def test_case_1_producing_nothing_is_reported_as_output_not_found(self):
        r = self._score(GOOD_SOLUTION, case1_from_solution=False)
        self.assertEqual(r["fail_reason"], "output-not-found")
        # The other two cases still run: only case 1 depends on the agent.
        self.assertEqual((r["n_pass"], r["n_cases"]), (2, 3))


# ------------------------------------------------------------------ end-to-end
@needs_openpyxl
class RunTaskTest(unittest.TestCase):
    def _run(self, responses, memory_block="", **cfg):
        d = tempfile.mkdtemp()
        self.addCleanup(lambda: __import__("shutil").rmtree(d, ignore_errors=True))
        make_task_dir(d, "13-1")
        client = FakeClient(responses)
        agent = SpreadsheetAgent(client, "fake", AgentConfig(max_turns=cfg.pop("max_turns", 8),
                                                            **cfg))
        ep = run_task(agent, TASK, d, os.path.join(d, "preds"), memory_block=memory_block)
        return ep, client

    def test_a_generalising_agent_succeeds_on_all_three_cases(self):
        ep, _ = self._run(solve_with(GOOD_SOLUTION))
        r = ep.rollouts[0]
        self.assertTrue(r.success)
        self.assertEqual(r.reward, 1.0)
        self.assertEqual((r.meta["n_pass"], r.meta["n_cases"]), (3, 3))
        self.assertTrue(r.meta["wrote_solution"])
        self.assertEqual(ep.scope, {"env": "spreadsheetbench", "task_type": "cell_level"})

    def test_a_hardcoding_agent_gets_partial_reward_but_not_success(self):
        ep, _ = self._run(solve_with(HARDCODED_SOLUTION))
        r = ep.rollouts[0]
        self.assertFalse(r.success)
        self.assertAlmostEqual(r.reward, 1 / 3)

    def test_the_dataset_workbook_is_never_written_through(self):
        d = tempfile.mkdtemp()
        self.addCleanup(lambda: __import__("shutil").rmtree(d, ignore_errors=True))
        task_dir = make_task_dir(d, "13-1")
        original = Path(os.path.join(task_dir, "1_13-1_input.xlsx")).read_bytes()
        agent = SpreadsheetAgent(FakeClient(solve_with(GOOD_SOLUTION)), "fake",
                                 AgentConfig(max_turns=8))
        run_task(agent, TASK, d, os.path.join(d, "preds"))
        # Corrupting the dataset would only show up as unexplained score drift
        # in later episodes, so this is worth asserting directly.
        self.assertEqual(Path(os.path.join(task_dir, "1_13-1_input.xlsx")).read_bytes(), original)

    def test_auto_verify_warns_about_a_formula_with_no_cached_value(self):
        ep, _ = self._run(solve_with(FORMULA_SOLUTION))
        observations = "\n".join(s.observation for s in ep.rollouts[0].steps)
        self.assertIn("formula with NO cached value", observations)

    def test_auto_verify_never_reveals_the_expected_answer(self):
        # SkillOpt has a second verifier that diffs against the golden workbook.
        # Wiring that in here would leak the answer into the trajectory and from
        # there into every memory entry written from it.
        ep, _ = self._run(solve_with(GOOD_SOLUTION))
        observations = "\n".join(s.observation for s in ep.rollouts[0].steps)
        for word in ("expected", "golden", "gt="):
            self.assertNotIn(word, observations)

    def test_memory_block_reaches_the_system_prompt(self):
        _, client = self._run(solve_with(GOOD_SOLUTION), memory_block="- reopen and check")
        self.assertIn("- reopen and check", client.calls[0][0]["content"])

    def test_injected_memory_tokens_are_recorded(self):
        ep, _ = self._run(solve_with(GOOD_SOLUTION), memory_block="- reopen and check")
        self.assertGreater(ep.rollouts[0].meta["injected_memory_tokens"], 0)

    def test_a_missing_block_is_a_nudge_before_it_is_a_stop(self):
        # Ending on the first formatting slip would score an episode the agent
        # never attempted.
        ep, client = self._run(["I'll think about it."] + solve_with(GOOD_SOLUTION))
        self.assertTrue(ep.rollouts[0].success)
        self.assertEqual(ep.rollouts[0].meta["parse_failures"], 1)

    def test_prose_only_agent_stops_without_crashing(self):
        ep, _ = self._run(["thinking", "still thinking", "done"])
        r = ep.rollouts[0]
        self.assertFalse(r.success)
        self.assertEqual(r.reward, 0.0)
        self.assertFalse(r.meta["wrote_solution"])

    def test_turn_limit_is_recorded_as_the_error(self):
        ep, _ = self._run(["```bash\necho hi\n```"] * 6, max_turns=3)
        self.assertEqual(ep.rollouts[0].error, "turn_limit_reached")
        self.assertEqual(len(ep.rollouts[0].steps), 3)

    def test_stop_sequence_is_passed_to_the_server(self):
        _, client = self._run(solve_with(GOOD_SOLUTION))
        self.assertIn("\nOutput:\n```", client.last_stop)


# ------------------------------------------------------------------ manifests
class ManifestTest(unittest.TestCase):
    def test_scope_key_follows_the_instruction_type(self):
        self.assertEqual(TASK.scope()["task_type"], "cell_level")
        sheet = TaskSpec("x", "test", instruction_type="Sheet-Level Manipulation")
        self.assertEqual(sheet.scope()["task_type"], "sheet_level")
        self.assertEqual(TaskSpec("x", "test").scope()["task_type"], "other")

    def test_answer_position_is_qualified_with_the_answer_sheet(self):
        self.assertEqual(TASK.eval_answer_position, "Sheet1!B1")

    def test_an_already_qualified_answer_position_is_left_alone(self):
        t = TaskSpec("x", "test", answer_position="Other!A1", answer_sheet="Sheet1")
        self.assertEqual(t.eval_answer_position, "Other!A1")

    def test_shipped_manifests_load_and_are_disjoint(self):
        root = Path(__file__).resolve().parent.parent / "manifests"
        evolve = root / "spreadsheetbench_evolve_train_50_seed42.json"
        ev = root / "spreadsheetbench_eval_test_100_seed42.json"
        if not evolve.is_file() or not ev.is_file():
            self.skipTest("manifests not built")
        a, b = load_manifest(evolve), load_manifest(ev)
        self.assertEqual((len(a), len(b)), (50, 100))
        self.assertFalse({t.task_id for t in a} & {t.task_id for t in b})
        # Tasks sharing the leading number are questions over the same workbook;
        # splitting a group across the two phases would leak.
        groups = lambda ts: {t.task_id.split("-")[0] for t in ts}  # noqa: E731
        self.assertFalse(groups(a) & groups(b))
        self.assertTrue(all(t.instruction and t.answer_position for t in a + b))


if __name__ == "__main__":
    unittest.main()
