"""SpreadsheetBench adapter: a tool-using ReAct agent that produces `Episode` objects.

Same contract as the ALFWorld, WebShop and AppWorld adapters -- TaskSpec/manifest
in, `Episode` out -- so the memory systems are unchanged. Four things differ, and
each one changes how the results must be read.

**The agent edits a real workbook through a shell.** Each turn it emits one
fenced block: a ```python block is written to `solution.py`, a ```bash block is
executed in a scratch directory. There is no environment server and no episode
"action space" -- the world is a temp directory holding a copy of the task's
input `.xlsx`, and the agent's effect on it is whatever its code did. Nothing
scores until the agent stops.

**Success is defined across three test cases, not one.** A SpreadsheetBench task
ships `1_<id>_input.xlsx` .. `3_<id>_input.xlsx` with matching `_answer` files.
The agent only ever sees case 1; afterwards its `solution.py` is re-executed
against cases 2 and 3 with the paths swapped. This is the benchmark's whole
anti-hardcoding device: an agent that reads the preview and writes literal
answers passes case 1 and fails the rest. `success` (all cases pass) is the
strict number and `reward` (`n_pass / n_cases`) is the graded one, so as on
WebShop and AppWorld the two can move apart and both are reported.

**The evaluator is a port, and it is unforgiving.** `compare_workbooks` below
follows upstream's `evaluation/evaluation.py` (RUCKBReasoning/SpreadsheetBench):
numbers are compared after `round(float(v), 2)` -- a quantization, not a
tolerance -- `None` and `""` are equal, and otherwise a type mismatch fails the
cell outright. Formatting is not compared, because upstream does not compare it
either. The single most common way to score 0 with correct arithmetic is to
write an Excel *formula*: openpyxl does not evaluate formulas, so the graded
cell reads back as `None`. `CRITICAL_RULES` warns about this and `_auto_verify`
detects it after every `python solution.py`, both carried over from SkillOpt's
scaffold, which is where that failure mode was diagnosed.

**Scaffold provenance: written here, like WebShop's.** The upstream and SkillOpt
ReAct agents both drive the model through OpenAI *native tool calls*
(`bash`, `write_file`). The shared vLLM server in `scripts/serve_qwen.sh` is not
launched with `--enable-auto-tool-choice`, and turning that on would change the
server every other benchmark in this study is already running against. So the
same two tools are expressed as fenced blocks and parsed here, in the same style
as the other three adapters. `CRITICAL_RULES` and the protocol are semantically
SkillOpt's; the wording is not byte-identical to anything published. Treat the
absolute number as uncalibrated and read the `none` arm as the only reference,
exactly as on WebShop.

**The agent runs arbitrary shell commands.** `_run_bash` is `subprocess.run(...,
shell=True)` with a timeout and `cwd` set to a per-episode temp directory, which
is what upstream does. It is not a sandbox: nothing stops a model from touching
files outside that directory. Run sweeps as an unprivileged user on a scratch
filesystem.
"""

from __future__ import annotations

import datetime
import glob as _glob
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import textwrap
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Sequence

from ..episode import Episode, Rollout, Step
from ..tokens import count_tokens

SCOPE_ENV = "spreadsheetbench"


# ======================================================= official evaluator
# Port of RUCKBReasoning/SpreadsheetBench `evaluation/evaluation.py`, by way of
# SkillOpt's `skillopt/envs/spreadsheetbench/evaluator.py`. Kept verbatim in
# behaviour: any "improvement" here silently redefines the benchmark.

def _datetime_to_float(dt: datetime.datetime) -> float:
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def _transform_value(v):
    if isinstance(v, bool):
        # Upstream does not special-case bools, but `round(float(True), 2) == 1.0`
        # is what its int/float branch would produce anyway; promoting explicitly
        # keeps `1` and `True` comparing equal rather than failing the type check.
        return round(float(v), 2)
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, datetime.time):
        return str(v)[:-3]
    if isinstance(v, datetime.datetime):
        return round(_datetime_to_float(v), 0)
    if isinstance(v, str):
        try:
            return round(float(v), 2)
        except ValueError:
            return v
    return v


def compare_cell_value(v1, v2) -> bool:
    v1 = _transform_value(v1)
    v2 = _transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) is not type(v2):
        return False
    return v1 == v2


def _col_num2name(n: int) -> str:
    name = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        name = chr(65 + r) + name
    return name


def _col_name2num(name: str) -> int:
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


def _parse_range(range_str: str):
    start_cell, end_cell = range_str.split(":")
    sc = "".join(ch for ch in start_cell if ch.isalpha())
    sr = "".join(ch for ch in start_cell if ch.isdigit())
    ec = "".join(ch for ch in end_cell if ch.isalpha())
    er = "".join(ch for ch in end_cell if ch.isdigit())
    return (_col_name2num(sc), int(sr)), (_col_name2num(ec), int(er))


def generate_cell_names(range_str: str) -> list[str]:
    if ":" not in range_str:
        return [range_str]
    (sc, sr), (ec, er) = _parse_range(range_str)
    cols = [_col_num2name(i) for i in range(sc, ec + 1)]
    return [f"{c}{r}" for c in cols for r in range(sr, er + 1)]


def iter_answer_ranges(answer_position: str, default_sheet: str):
    """Yield `(sheet_name, cell_range)` for each comma-separated range."""
    for scr in (answer_position or "").split(","):
        scr = scr.strip()
        if not scr:
            continue
        if "!" in scr:
            sheet_name, cell_range = scr.split("!", 1)
            sheet_name = sheet_name.strip().strip("'\"")
        else:
            sheet_name, cell_range = default_sheet, scr
        yield sheet_name, cell_range.strip().strip("'\"")


def compare_workbooks(gt_file: str, proc_file: str, answer_position: str) -> tuple[bool, str, float]:
    """One test case. Returns `(all_cells_match, first_mismatch, cell_match_rate)`.

    `cell_match_rate` is not part of upstream's protocol and never contributes to
    a reported score. It exists because the pass/fail bool cannot distinguish "one
    cell off by a rounding rule" from "wrote nothing at all", and that distinction
    is what makes a failed episode readable in `RESULTS`.
    """
    import openpyxl

    if not os.path.exists(proc_file):
        return False, "file not exist", 0.0
    try:
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return False, f"load error: {exc}", 0.0
    try:
        ok_all, msg_first = True, ""
        n_cells = n_match = 0
        for sheet_name, cell_range in iter_answer_ranges(answer_position, wb_gt.sheetnames[0]):
            if sheet_name not in wb_proc.sheetnames:
                ok_all = False
                msg_first = msg_first or f"worksheet not found: {sheet_name}"
                n_cells += len(generate_cell_names(cell_range))
                continue
            ws_gt, ws_proc = wb_gt[sheet_name], wb_proc[sheet_name]
            for cn in generate_cell_names(cell_range):
                n_cells += 1
                gv, pv = ws_gt[cn].value, ws_proc[cn].value
                if compare_cell_value(gv, pv):
                    n_match += 1
                    continue
                ok_all = False
                msg_first = msg_first or f"value@{sheet_name}!{cn}: gt={gv!r} pred={pv!r}"
        return ok_all, msg_first, (n_match / n_cells if n_cells else 0.0)
    finally:
        wb_gt.close()
        wb_proc.close()


# ================================================== generated-code executor
# Port of SkillOpt's `executor.py`: re-run the agent's `solution.py` against a
# different test case by overriding the two paths it declares at the top.

_RUNNER_TEMPLATE = textwrap.dedent(
    """
    import os, sys, traceback
    INPUT_PATH = {input_path!r}
    OUTPUT_PATH = {output_path!r}
    try:
    {user_code_indented}
    except Exception:
        traceback.print_exc()
        sys.exit(2)
    """
)

_PATH_ASSIGN_RE = re.compile(r"^\s*(INPUT_PATH|OUTPUT_PATH)\s*=\s*.+$", re.MULTILINE)


def run_generated_code(code: str, input_path: str, output_path: str,
                       timeout: float = 180.0) -> tuple[bool, str]:
    """Execute `code` with `INPUT_PATH`/`OUTPUT_PATH` forced to the given paths.

    The agent's own assignments are stripped rather than trusted: on case 2 they
    still point at case 1's files, and a solution that ignored the override would
    re-score case 1 three times and look like it generalised.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    cleaned = _PATH_ASSIGN_RE.sub("", code)
    script = _RUNNER_TEMPLATE.format(
        input_path=input_path,
        output_path=output_path,
        user_code_indented=textwrap.indent(cleaned, "    "),
    )
    with tempfile.NamedTemporaryFile("w", suffix=".py", delete=False) as fh:
        fh.write(script)
        tmp = fh.name
    try:
        proc = subprocess.run([sys.executable, tmp], capture_output=True, text=True,
                              timeout=timeout if timeout and timeout > 0 else None)
        if proc.returncode != 0:
            return False, (proc.stdout + "\n" + proc.stderr).strip()
        if not os.path.exists(output_path):
            return False, "output file was not created"
        return True, ""
    except subprocess.TimeoutExpired:
        return False, f"timeout after {timeout}s"
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass


# ============================================================ task manifests
@dataclass(frozen=True)
class TaskSpec:
    task_id: str
    split: str
    instruction: str = ""
    #: "Cell-Level Manipulation" | "Sheet-Level Manipulation", verbatim from
    #: `dataset.json`. Shown to the agent, as upstream shows it.
    instruction_type: str = ""
    #: Directory holding the test-case workbooks, relative to `--data-root`.
    spreadsheet_path: str = ""
    #: The graded cell range, e.g. `"A3:D32"` or `"Sheet1!B2,Sheet1!D5"`.
    #: This is task metadata, not ground truth -- upstream's own agent is given
    #: it, and without it the task is underspecified.
    answer_position: str = ""
    answer_sheet: str = ""
    #: Where the input data sits. Recorded for completeness; not shown.
    data_position: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "task_id": self.task_id,
            "split": self.split,
            "instruction": self.instruction,
            "instruction_type": self.instruction_type,
            "spreadsheet_path": self.spreadsheet_path,
            "answer_position": self.answer_position,
            "answer_sheet": self.answer_sheet,
            "data_position": self.data_position,
        }

    @classmethod
    def from_dict(cls, d: dict[str, Any], default_split: str = "train") -> "TaskSpec":
        task_id = str(d.get("task_id") or d.get("id") or "")
        if not task_id:
            raise ValueError(f"manifest entry has no task_id: {d}")
        return cls(
            task_id=task_id,
            split=str(d.get("split") or default_split),
            instruction=str(d.get("instruction") or ""),
            instruction_type=str(d.get("instruction_type") or ""),
            spreadsheet_path=str(d.get("spreadsheet_path") or f"spreadsheet/{task_id}"),
            answer_position=str(d.get("answer_position") or ""),
            answer_sheet=str(d.get("answer_sheet") or ""),
            data_position=str(d.get("data_position") or ""),
        )

    @property
    def task_type(self) -> str:
        """The memsys `task_type` scope key.

        Only two values, because SpreadsheetBench only labels two. That is a
        coarser grouping than ALFWorld's six procedure families -- closer to
        WebShop's product departments -- so expect little from batch induction
        clustering on it, and do not read that as a property of a memory type.
        """
        low = (self.instruction_type or "").lower()
        if "cell" in low:
            return "cell_level"
        if "sheet" in low:
            return "sheet_level"
        return "other"

    @property
    def eval_answer_position(self) -> str:
        """`answer_position` qualified with `answer_sheet` when it carries no sheet."""
        if self.answer_position and self.answer_sheet and "!" not in self.answer_position:
            return f"{self.answer_sheet}!{self.answer_position}"
        return self.answer_position

    def scope(self) -> dict[str, str]:
        return {"env": SCOPE_ENV, "task_type": self.task_type}


def load_manifest(path: str | Path) -> list[TaskSpec]:
    p = Path(path)
    value = json.loads(p.read_text(encoding="utf-8"))
    if isinstance(value, list):
        return [TaskSpec.from_dict(item) for item in value]
    if not isinstance(value, dict) or not isinstance(value.get("tasks"), list):
        raise ValueError(f"unsupported manifest format: {p}")
    split = str(value.get("split") or "train")
    return [TaskSpec.from_dict(item, default_split=split) for item in value["tasks"]]


# ========================================================= test-case lookup
def find_test_cases(task_dir: str | Path) -> list[tuple[str, str, str]]:
    """Return `[(case_no, input_path, answer_path), ...]` sorted by case number.

    Three naming conventions are in circulation and the adapter must read all of
    them, because which one a task uses depends on the archive it came from:

      * ``{no}_{id}_input.xlsx`` + ``{no}_{id}_answer.xlsx``  -- `all_data_912_v0.1`
      * ``{no}_{id}_init.xlsx``  + ``{no}_{id}_golden.xlsx``  -- `verified_400`
      * ``initial.xlsx``         + ``golden.xlsx``            -- `verified_400`, 5 tasks

    Only the first form ships three cases per task. On `verified_400` every task
    has exactly one, which collapses `reward` onto `success` and removes the
    generalisation check entirely -- see the module docstring and RUN_SPREADSHEETBENCH.md.
    """
    task_dir = str(task_dir)
    cases: list[tuple[str, str, str]] = []
    for suffix, answer_suffix in (("_input.xlsx", "_answer.xlsx"), ("_init.xlsx", "_golden.xlsx")):
        for ip in sorted(_glob.glob(os.path.join(task_dir, f"*{suffix}"))):
            ap = ip[: -len(suffix)] + answer_suffix
            if os.path.exists(ap):
                cases.append((os.path.basename(ip).split("_", 1)[0], ip, ap))
    if not cases:
        bare_init = os.path.join(task_dir, "initial.xlsx")
        bare_gold = os.path.join(task_dir, "golden.xlsx")
        if os.path.exists(bare_init) and os.path.exists(bare_gold):
            cases.append(("1", bare_init, bare_gold))
    return sorted(cases, key=lambda c: (len(c[0]), c[0]))


# =================================================================== prompt
#: Verbatim in substance from SkillOpt's `prompts/critical_rules.md`. Rule 1 is
#: not style advice: it is the difference between a correct computation scoring
#: 1.0 and scoring 0.0, because the evaluator reads cached values and openpyxl
#: writes none.
CRITICAL_RULES = """## Critical Rules (MUST follow)
1. NEVER write Excel formulas into cells that will be graded on their displayed value.
   openpyxl does NOT compute formulas -- the evaluator reads the cached value and sees None.
   Compute the result in Python and write a literal number or string.
2. After saving the workbook, reopen it and print the graded cells to check what was written:
   `wb2 = openpyxl.load_workbook(OUTPUT_PATH, data_only=True); print(wb2[sheet][cell].value)`
3. Do NOT hardcode values read from a preview. Your solution.py is re-executed against
   other input files with the same structure but different data, and all of them must pass.
4. Use only the standard library, openpyxl and pandas. `pandas.to_excel` destroys existing
   formulas, formatting and extra sheets -- write back with openpyxl when they must survive."""

SYSTEM_PROMPT_TEMPLATE = """You are an expert spreadsheet manipulation agent. You solve a task by writing a Python script and running it in a shell.

{critical_rules}
{memory_section}## Protocol
Each of your replies contains a short plan followed by exactly ONE fenced block.

To write a file, emit a python block (it is saved as solution.py and NOT executed):

```python
INPUT_PATH  = "<the exact input path given in the task>"
OUTPUT_PATH = "<the exact output path given in the task>"
import openpyxl
...
wb.save(OUTPUT_PATH)
```

To run a command, emit a bash block (executed in the working directory; you get stdout+stderr):

```bash
python solution.py
```

solution.py MUST define INPUT_PATH and OUTPUT_PATH on its first two lines, exactly as
given in the task, and MUST save the result to OUTPUT_PATH.

Suggested order of work:
1. A bash block that inspects the input workbook (sheet names, headers, dimensions, a few rows).
2. A python block that writes solution.py.
3. A bash block running `python solution.py`, then read the verification report.
4. Repeat 2-3 until the graded cells hold the right values.

When the output file is written and correct, reply with the single word DONE and no fenced block."""

#: What the `none` arm puts where the memory block goes. The section cannot just
#: be dropped: removing it would make the no-memory arm a different scaffold
#: rather than the same scaffold with an empty memory.
EMPTY_MEMORY = ("## Archived memory\n(No entries yet -- rely on the rules and protocol above.)\n\n")

#: Same wording as the ALFWorld and WebShop adapters, so the injected block is
#: framed identically in every benchmark and a difference in effect is not a
#: difference in framing.
MEMORY_HEADER = (
    "**Archived memory (distilled from your own past episodes in this environment):**"
)


def build_system_prompt(memory_block: str) -> str:
    if memory_block.strip():
        memory_section = f"## Archived memory\n{MEMORY_HEADER}\n{memory_block.strip()}\n\n"
    else:
        memory_section = EMPTY_MEMORY
    return SYSTEM_PROMPT_TEMPLATE.format(
        critical_rules=CRITICAL_RULES, memory_section=memory_section
    )


def build_user_prompt(task: TaskSpec, input_path: str, output_path: str) -> str:
    parts = [
        f"# Instruction\n{task.instruction}",
        f"# Input file\n{input_path}",
        f"# Output file\n{output_path}",
    ]
    if task.instruction_type:
        parts.append(f"# Instruction type\n{task.instruction_type}")
    if task.eval_answer_position:
        parts.append(f"# Answer position\n{task.eval_answer_position}")
    parts.append(
        "Manipulate the input spreadsheet according to the instruction and save the "
        "result to the output file."
    )
    return "\n\n".join(parts)


#: `(language, optional path, body)`. The first block in a reply wins, for the
#: same reason the AppWorld adapter takes the first code block: a model that
#: emits several has invented the outputs between them, so every block after the
#: first is written against a shell state that never existed.
_BLOCK_RE = re.compile(r"```(python|bash|sh|shell)[ \t]*([^\n`]*)\n(.*?)```", re.S)

#: Stop before the model can write the shell's reply and keep going. Only the
#: observation fence is listed: anything resembling task-prompt text would also
#: match a comment inside a legitimate code block and truncate it mid-script.
STOP_SEQUENCES = ["\nOutput:\n```"]


def parse_action(text: str) -> tuple[str, str, str] | None:
    """Return `(kind, path, body)` with `kind` in `{"write", "bash"}`, or None."""
    m = _BLOCK_RE.search(str(text or ""))
    if not m:
        return None
    lang, path, body = m.group(1), m.group(2).strip(), m.group(3)
    if lang == "python":
        return "write", (path or "solution.py"), body
    return "bash", "", body.strip()


# ================================================================ workspace
def _auto_verify(work_dir: str, output_path: str) -> str:
    """Inspect the produced workbook after `python solution.py`.

    Carried over from SkillOpt's ReAct scaffold, and deliberately **gold-free**:
    it reports what the agent wrote, never what the answer is. SkillOpt has a
    second verifier that diffs against the golden workbook, but that one is only
    wired into its code-generation path; putting it here would leak the answer
    into the trajectory, and from there into every memory entry written from it.

    The formula-with-no-cached-value check is the whole point. It is the single
    most common way to lose a task whose arithmetic was right.
    """
    import openpyxl

    path = output_path
    if not path or not os.path.exists(path):
        candidates = [f for f in _glob.glob(os.path.join(work_dir, "*.xlsx"))
                      if "_pred" in os.path.basename(f)]
        path = candidates[0] if candidates else ""
    if not path or not os.path.exists(path):
        return ("\n\n[VERIFY] Output file not found. Check that OUTPUT_PATH is the exact "
                "path given in the task and that wb.save(OUTPUT_PATH) ran.")

    try:
        wb_f = openpyxl.load_workbook(path, data_only=False)
        wb_v = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return f"\n\n[VERIFY] Could not open the output workbook: {exc}"
    try:
        lines = [f"\n\n[VERIFY] Output file exists: {path}"]
        sn = wb_f.sheetnames[0]
        ws_f, ws_v = wb_f[sn], wb_v[sn]
        lines.append(f"  Sheets: {wb_f.sheetnames}")
        lines.append(f"  Sheet '{sn}': {ws_f.dimensions}")
        for row in ws_v.iter_rows(min_row=1, max_row=min(5, ws_v.max_row), values_only=True):
            lines.append(f"    {list(row)}")

        stale: list[str] = []
        for row_f, row_v in zip(
            ws_f.iter_rows(min_row=1, max_row=min(30, ws_f.max_row)),
            ws_v.iter_rows(min_row=1, max_row=min(30, ws_v.max_row)),
        ):
            for cf, cv in zip(row_f, row_v):
                if isinstance(cf.value, str) and cf.value.startswith("=") and cv.value is None:
                    stale.append(cf.coordinate)
        if stale:
            lines.append(f"  WARNING: {len(stale)} cells hold a formula with NO cached value -- "
                         f"the evaluator will read None: {stale[:10]}")
            lines.append("  FIX: compute the value in Python and write a literal instead.")
        else:
            lines.append("  All inspected cells hold concrete values.")
        return "\n".join(lines)
    except Exception as exc:  # noqa: BLE001
        return f"\n\n[VERIFY] Could not inspect the output: {exc}"
    finally:
        wb_f.close()
        wb_v.close()


def _shell_env() -> dict[str, str]:
    """Environment for the agent's shell, with this interpreter first on PATH.

    Without this, a bare `python solution.py` resolves to whatever `python` the
    login shell happens to expose, which on a cluster is rarely the environment
    memsys is running in. The failure is total but reads like a model error:
    every episode ends in `ModuleNotFoundError: No module named 'openpyxl'`
    against a prompt that told the agent openpyxl was available.

    It also has to match `run_generated_code`, which re-runs `solution.py` with
    `sys.executable` at scoring time. If the two interpreters differ, a script
    can work in the loop and fail on cases 2 and 3 for reasons the trajectory
    never shows.
    """
    env = os.environ.copy()
    bindir = os.path.dirname(sys.executable)
    env["PATH"] = bindir + os.pathsep + env.get("PATH", "")
    return env


def _run_bash(cmd: str, work_dir: str, output_path: str, timeout: float = 120.0,
              max_chars: int = 4000) -> str:
    try:
        proc = subprocess.run(cmd, shell=True, capture_output=True, text=True,
                              timeout=timeout, cwd=work_dir, env=_shell_env())
        out = (proc.stdout + proc.stderr).strip()
    except subprocess.TimeoutExpired:
        return f"[timeout after {timeout:g}s]"
    except Exception as exc:  # noqa: BLE001
        return f"[error: {exc}]"
    if len(out) > max_chars:
        out = out[: max_chars - 200] + f"\n...[truncated, {len(out)} chars total]"
    result = out or "(no output)"
    if "solution.py" in cmd and "python" in cmd.lower():
        result += _auto_verify(work_dir, output_path)
    return result


def _write_file(path: str, content: str, work_dir: str) -> str:
    try:
        full = path if os.path.isabs(path) else os.path.join(work_dir, path)
        parent = os.path.dirname(full)
        if parent:
            os.makedirs(parent, exist_ok=True)
        with open(full, "w", encoding="utf-8") as fh:
            fh.write(content)
        n_lines = content.count("\n") + 1
        return f"File written: {full} ({len(content)} chars, {n_lines} lines)"
    except Exception as exc:  # noqa: BLE001
        return f"[write_file error: {exc}]"


# ==================================================================== agent
@dataclass
class AgentConfig:
    #: Upstream and SkillOpt both use 30. Kept, because the loop here is
    #: inspect -> write -> run -> fix and each repair costs two turns.
    max_turns: int = 30
    #: Far above ALFWorld/WebShop's 256: a turn must fit a whole solution.py.
    max_tokens: int = 1536
    temperature: float = 0.7  # Qwen3.5 non-thinking recommendation
    #: The server is served with --max-model-len 32768.
    context_limit_tokens: int = 26000
    min_history_steps: int = 3
    #: One `print(df)` over a large sheet can return hundreds of KB.
    max_output_chars: int = 4000
    #: Per-command wall clock inside the agent loop.
    bash_timeout: float = 120.0
    #: How much of a written file is kept in `Step.action`. The file itself is
    #: written in full; this only bounds what the memory writers read, and an
    #: untruncated 300-line solution.py would consume a whole rollout budget.
    max_recorded_file_chars: int = 2000


class SpreadsheetAgent:
    """ReAct agent over a scratch directory holding one task's input workbook."""

    def __init__(self, client: Any, model: str, config: AgentConfig | None = None):
        self.client = client
        self.model = model
        self.config = config or AgentConfig()

    def _trim(self, messages: list[dict[str, str]], head: int) -> list[dict[str, str]]:
        total = sum(count_tokens(m["content"]) for m in messages)
        while total > self.config.context_limit_tokens and (
            len(messages) - head > 2 * self.config.min_history_steps
        ):
            dropped = messages[head : head + 2]
            del messages[head : head + 2]
            total -= sum(count_tokens(m["content"]) for m in dropped)
        return messages

    def _observation_turn(self, obs: str) -> str:
        text = obs.strip()
        if len(text) > self.config.max_output_chars:
            text = text[: self.config.max_output_chars] + "\n...<truncated>"
        return f"Output:\n```\n{text}\n```"

    def run(
        self,
        task: TaskSpec,
        work_dir: str,
        input_path: str,
        output_path: str,
        memory_block: str = "",
        rollout_id: str = "r0",
        temperature: float | None = None,
    ) -> tuple[list[Step], dict[str, Any]]:
        """Drive the loop. Returns `(steps, meta)`; scoring happens in `run_task`."""
        messages = [
            {"role": "system", "content": build_system_prompt(memory_block)},
            {"role": "user", "content": build_user_prompt(task, input_path, output_path)},
        ]
        head = len(messages)

        steps: list[Step] = []
        error = None
        parse_failures = bash_errors = 0
        prompt_tokens = completion_tokens = 0
        wrote_solution = nudged = False

        for _ in range(self.config.max_turns):
            self._trim(messages, head)
            try:
                resp = self.client.chat.completions.create(
                    model=self.model,
                    messages=messages,
                    temperature=self.config.temperature if temperature is None else temperature,
                    max_tokens=self.config.max_tokens,
                    stop=STOP_SEQUENCES,
                )
            except Exception as exc:  # noqa: BLE001
                error = f"llm_error: {type(exc).__name__}: {exc}"
                break
            usage = getattr(resp, "usage", None)
            prompt_tokens += getattr(usage, "prompt_tokens", 0) or 0
            completion_tokens += getattr(usage, "completion_tokens", 0) or 0
            text = resp.choices[0].message.content or ""

            action = parse_action(text)
            if action is None:
                # No block. Upstream reads this as "the agent is finished", and so
                # do we -- but only once it has actually produced a solution.py.
                # Before that it is far more likely to be a formatting slip, and
                # ending there would score an episode the agent never attempted.
                if wrote_solution or nudged:
                    steps.append(Step(action="(stop)", observation="", thought=text[:400]))
                    break
                parse_failures += 1
                nudged = True  # re-prompt at most once, then take the next stop at face value
                messages.append({"role": "assistant", "content": text})
                messages.append({"role": "user", "content":
                                 "Output:\n```\nNo fenced block found. Reply with exactly one "
                                 "```python block (saved as solution.py) or one ```bash block "
                                 "(executed), or DONE if the output file is already correct.\n```"})
                steps.append(Step(action="", observation="no fenced block", thought=text[:400]))
                continue

            kind, path, body = action
            if kind == "write":
                obs = _write_file(path, body, work_dir)
                if os.path.basename(path) == "solution.py":
                    wrote_solution = True
                recorded = body
                if len(recorded) > self.config.max_recorded_file_chars:
                    recorded = recorded[: self.config.max_recorded_file_chars] + "\n...<truncated>"
                steps.append(Step(action=f"write_file {path}\n{recorded}", observation=obs,
                                  thought=text.split("```")[0].strip()[:800]))
            else:
                obs = _run_bash(body, work_dir, output_path, timeout=self.config.bash_timeout,
                                max_chars=self.config.max_output_chars)
                if "Traceback (most recent call last)" in obs or obs.startswith("[error"):
                    bash_errors += 1
                steps.append(Step(action=f"bash {body}", observation=obs,
                                  thought=text.split("```")[0].strip()[:800]))

            messages.append({"role": "assistant", "content": text})
            messages.append({"role": "user", "content": self._observation_turn(obs)})
        else:
            error = error or "turn_limit_reached"

        return steps, {
            "error": error,
            "parse_failures": parse_failures,
            "bash_errors": bash_errors,
            "wrote_solution": wrote_solution,
            "prompt_tokens": prompt_tokens,
            "completion_tokens": completion_tokens,
            "injected_memory_tokens": count_tokens(memory_block),
        }


# ================================================================= episodes
def score_task(
    task: TaskSpec,
    cases: Sequence[tuple[str, str, str]],
    out_dir: str,
    solution_path: str,
    exec_timeout: float = 180.0,
) -> dict[str, Any]:
    """Grade every test case. Case 1 uses whatever the agent produced; the rest
    re-execute its `solution.py` against a fresh input.

    An agent that never wrote a `solution.py` is not disqualified: if it edited
    case 1's workbook by hand it still scores that case, and failing the other
    two is the correct outcome rather than an error.
    """
    result: dict[str, Any] = {
        "n_cases": len(cases), "n_exec_ok": 0, "n_pass": 0,
        "cases": [], "fail_reason": "", "cell_match_rate": 0.0,
    }
    code = ""
    if os.path.exists(solution_path):
        code = Path(solution_path).read_text(encoding="utf-8", errors="replace")

    rates: list[float] = []
    for i, (no, input_path, answer_path) in enumerate(cases):
        pred_path = os.path.join(out_dir, f"{no}_pred.xlsx")
        if i > 0:
            if not code:
                result["cases"].append({"no": no, "stage": "exec", "ok": False,
                                        "error": "no-solution-py"})
                result["fail_reason"] = result["fail_reason"] or "no-solution-py-for-other-cases"
                rates.append(0.0)
                continue
            ok_exec, err = run_generated_code(code, input_path, pred_path, timeout=exec_timeout)
            if not ok_exec:
                tail = err.strip().splitlines()[-1][:200] if err.strip() else "unknown"
                result["cases"].append({"no": no, "stage": "exec", "ok": False, "error": err[:500]})
                result["fail_reason"] = result["fail_reason"] or f"exec-error: {tail}"
                rates.append(0.0)
                continue
        if not os.path.exists(pred_path):
            result["cases"].append({"no": no, "stage": "exec", "ok": False,
                                    "error": "output-not-found"})
            result["fail_reason"] = result["fail_reason"] or "output-not-found"
            rates.append(0.0)
            continue

        result["n_exec_ok"] += 1
        try:
            ok, msg, rate = compare_workbooks(answer_path, pred_path, task.eval_answer_position)
        except Exception as exc:  # noqa: BLE001
            ok, msg, rate = False, f"eval-exception: {type(exc).__name__}: {exc}", 0.0
        rates.append(rate)
        if ok:
            result["n_pass"] += 1
        else:
            result["fail_reason"] = result["fail_reason"] or f"eval-mismatch: {msg[:200]}"
        result["cases"].append({"no": no, "stage": "eval", "ok": ok, "reason": msg,
                                "cell_match_rate": round(rate, 4)})

    n = result["n_cases"]
    result["cell_match_rate"] = round(sum(rates) / len(rates), 4) if rates else 0.0
    # The graded number, as on WebShop and AppWorld.
    result["score"] = (result["n_pass"] / n) if n else 0.0
    # The strict one: SpreadsheetBench counts a task solved only if every test
    # case passes, which is what makes a hardcoded answer worthless.
    result["success"] = bool(n) and result["n_pass"] == n
    if result["success"]:
        result["fail_reason"] = ""
    return result


def run_task(
    agent: SpreadsheetAgent,
    task: TaskSpec,
    data_root: str | Path,
    out_dir: str | Path,
    memory_block: str = "",
    n_rollouts: int = 1,
    temperature: float | None = None,
    exec_timeout: float = 180.0,
    keep_workspace: bool = False,
) -> Episode:
    """Run the agent on one task and grade it. `out_dir` holds per-task artefacts."""
    sp = task.spreadsheet_path or f"spreadsheet/{task.task_id}"
    task_dir = sp if os.path.isabs(sp) else os.path.join(str(data_root), sp)
    cases = find_test_cases(task_dir)

    rollouts: list[Rollout] = []
    for i in range(n_rollouts):
        rollout_id = f"{task.task_id}#{i}"
        pred_dir = os.path.join(str(out_dir), task.task_id, str(i))
        os.makedirs(pred_dir, exist_ok=True)

        if not cases:
            rollouts.append(Rollout(rollout_id=rollout_id, steps=[], reward=0.0, success=False,
                                    error="no_test_cases",
                                    meta={"task_id": task.task_id, "task_family": task.task_type,
                                          "n_cases": 0, "fail_reason": "no-test-cases"}))
            continue

        no1, input1, _ = cases[0]
        output_path = os.path.join(pred_dir, f"{no1}_pred.xlsx")
        work_dir = tempfile.mkdtemp(prefix=f"ssb_{task.task_id}_{i}_")
        try:
            # The agent works on a copy. Writing through to the dataset would
            # corrupt the benchmark for every later episode, and the failure
            # would only surface as unexplained score drift.
            work_input = os.path.join(work_dir, os.path.basename(input1))
            shutil.copy2(input1, work_input)
            steps, meta = agent.run(task, work_dir, work_input, output_path,
                                    memory_block=memory_block, rollout_id=rollout_id,
                                    temperature=temperature)
            solution_src = os.path.join(work_dir, "solution.py")
            solution_dst = os.path.join(pred_dir, "solution.py")
            if os.path.exists(solution_src):
                shutil.copy2(solution_src, solution_dst)
        finally:
            if not keep_workspace:
                shutil.rmtree(work_dir, ignore_errors=True)

        scored = score_task(task, cases, pred_dir, os.path.join(pred_dir, "solution.py"),
                            exec_timeout=exec_timeout)
        rollouts.append(Rollout(
            rollout_id=rollout_id,
            steps=steps,
            reward=scored["score"],
            success=scored["success"],
            error=meta["error"],
            meta={
                "task_id": task.task_id,
                "task_family": task.task_type,
                "instruction_type": task.instruction_type,
                "answer_position": task.eval_answer_position,
                "n_steps": len(steps),
                "n_cases": scored["n_cases"],
                "n_pass": scored["n_pass"],
                "n_exec_ok": scored["n_exec_ok"],
                # Fraction of graded cells matching gold, averaged over cases.
                # Diagnostic only -- never a reported score. It separates "off by
                # one cell" from "wrote nothing", which pass/fail cannot.
                "cell_match_rate": scored["cell_match_rate"],
                "fail_reason": scored["fail_reason"],
                "cases": scored["cases"],
                **{k: meta[k] for k in ("parse_failures", "bash_errors", "wrote_solution",
                                        "prompt_tokens", "completion_tokens",
                                        "injected_memory_tokens")},
            },
        ))

    return Episode(
        task_id=task.task_id,
        instruction=task.instruction,
        rollouts=rollouts,
        scope=task.scope(),
        meta={"split": task.split, "instruction_type": task.instruction_type,
              "n_cases": len(cases)},
    )


def default_data_root() -> str:
    return os.environ.get(
        "SPREADSHEETBENCH_ROOT",
        "/gpfs/radev/scratch/cohan/jw3278/spreadsheetbench_root/all_data_912_v0.1",
    )
